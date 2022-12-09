from pypsexec.client import Client
import subprocess, re, os
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import requests
from bs4 import BeautifulSoup


userList = []
failedUsers = []
data = {}


def get_out(*args):
    return subprocess.check_output(args, shell=True).decode('866').rstrip().split('\n')[1]


def cmdParse(out):
    return out.decode('866').rstrip().replace('\r', '').split('\n')


def shellCommand(command):
    return subprocess.check_output(["powershell.exe", command]).decode('866').split('\n')


def output_rep(listik):
    longestKey = ''
    for i in listik:
        if len(i) > len(longestKey):
            longestKey = i

    for i in listik.items():
        print(' ' + (' ' * (len(longestKey) - len(i[0]))) + i[0] + ': ' + str(i[1]))


def getUserinfoList(user):
    usrtest = []
    page = requests.get(('http://cp/ad?search=' + user).replace(' ', '%20'), auth=('KiparenkoII', 'pass'))
    page.encoding = 'utf-8'

    username = re.search(r'\w+[?]', page.text).group(0).replace('?', '')

    page = requests.get('http://cp/ad/user/' + username, auth=('KiparenkoII', 'pass'))
    page.encoding = 'utf-8'

    soup = BeautifulSoup(page.text.replace('<br/>', 'IDSQ89NQ'), 'html.parser')
    ku4a = soup.find("div", id="main_content")

    content = ku4a.text.replace('IDSQ89NQ', '\n')

    for i in content.split('\n'):
        if len(i) > 0:
            usrtest.append(i)
    return usrtest


def sParse(command):
    while "  " in command:
        command = command.replace("  ", " ")
    return command


while True:
    upravlenie = sParse(input('Управление: ').replace('\"', ''))
    if not os.path.exists(upravlenie):
        os.mkdir(upravlenie)
    wb = load_workbook(filename='testo.xlsx')
    sheet = wb['Лист1']
    karet = ''
    amount = -1
    tmp = ''

    while karet != None:
        karet = sheet['A{}'.format(amount + 2)].value
        print(karet)
        amount += 1
    print('AMOUNT: ', amount)

    for i in sheet['A1:A{}'.format(amount)]:
        userList.append(i[0].value)

    for currentUser in userList:
        data.clear()
        wb = load_workbook(filename = 'ncard.xlsx') #, read_only=False)
        sheet = wb['Карточка АРМ']

        try:
            userInfo = getUserinfoList(currentUser)

            for cy, i in enumerate(userInfo):
                print(cy, '. ', i)

            data['user_fullname'] = currentUser

            sheet['B1'].value = data['user_fullname']

            data['position'] = userInfo[2].replace('Подразделение: ', '').replace('Администрация / ', '')
            data['department'] = userInfo[3].replace('Отдел: ', '')
            data['pc_name'] = re.findall(r'\w+$', userInfo[10].replace('\xa0', ''))[0].replace(' ', '')

            sheet['B6'].value = data['pc_name']

            if data['pc_name'] == 'да':
                failedUsers.append(currentUser)
                print('\n*Пользователь ' + currentUser + ' - неудачно')
                continue

        except:
            failedUsers.append(currentUser)
            print('\n*Пользователь ' + currentUser + ' - неудачно')
            continue

        try:
            c = Client(data['pc_name'], encrypt=False)
            c.connect()
            print('\n' + data['pc_name'] + ' connected')
            c.create_service()

            stdout, stderr, rc = c.run_executable("cmd.exe", arguments="/c wmic csproduct get name")
            data['pc_model'] = cmdParse(stdout)[1].replace('To Be Filled By O.E.M.', '').replace('System Product Name', '')

            if len(data['pc_model']) > 0:
                sheet['B3'].value = data['pc_model']
            else:
                data['pc_model'] = '-'
                sheet['B3'].value = data['pc_model']

            stdout, stderr, rc = c.run_executable("cmd.exe",arguments="/c arp -a -N " + data['pc_name'])
            data['ip'] = re.search(r'\d+[.]\d+[.]\d+[.]\d+', str(stdout)).group(0)

            stdout, stderr, rc = c.run_executable("cmd.exe",arguments="/c wmic baseboard get Manufacturer, product")
            data['baseboard'] = sParse(cmdParse(stdout)[1]).replace('To Be Filled By O.E.M.', '')
            sheet['D7'].value = data['baseboard']

            stdout, stderr, rc = c.run_executable("cmd.exe",arguments="/c wmic baseboard get serialnumber")
            tmp = sParse(cmdParse(stdout)[1]).replace('To be filled by O.E.M.', '')
            if len(tmp) > 0:
                data['baseboard'] += ' (s/n: ' + tmp + ')'

            stdout, stderr, rc = c.run_executable("cmd.exe",arguments="/c wmic CPU get Name")
            data['cpu'] = sParse(cmdParse(stdout)[1])
            sheet['D8'].value = data['cpu']

            stdout, stderr, rc = c.run_executable("cmd.exe",arguments="/c wmic computersystem get TotalPhysicalMemory")
            data['RAM'] = str(round((int(cmdParse(stdout)[1]) / 1024 / 1024 / 1024), 2)) + ' Gb'
            sheet['D9'].value = data['RAM']

            stdout, stderr, rc = c.run_executable("cmd.exe",arguments="/c wmic path win32_VideoController get name")
            data['graphics'] = ''
            for i in cmdParse(stdout)[1:]:
                if ('DameWare' not in i) and ('Mirror' not in i):
                    data['graphics'] += sParse(i) + ' '
            sheet['D10'].value = data['graphics']

            stdout, stderr, rc = c.run_executable("cmd.exe",arguments="/c wmic os get Caption")
            data['os'] = cmdParse(stdout)[1]

            if 'Windows 10' in data['os']:
                sheet['E12'] = '+'
            else:
                sheet['G12'] = '+'

            stdout, stderr, rc = c.run_executable("cmd.exe",arguments="/c wmic desktopmonitor get screenwidth, screenheight")
            for i in cmdParse(stdout)[1:]:
                if len(i) > 0:
                    tmp = sParse(i).split(' ')#.replace(' ', 'x')
            data['display_resolution'] = tmp[1] + 'x' + tmp[0]
            sheet['D18'].value = data['display_resolution']

            try:
                data['display_serial'] = shellCommand("get-wmiobject WmiMonitorID -computername " + data['pc_name'] + " -Namespace root\wmi | select serialnumberid | ForEach-Object {[System.Text.Encoding]::ASCII.GetString($_.serialnumberid)}")[0].replace('\x00', '').replace('\r', '')

                sheet['B15'].value = data['display_serial']
            except:
                pass

            try:
                data['display_name'] = shellCommand("get-wmiobject -computername " + data['pc_name'] + " WmiMonitorID -Namespace root\wmi | select UserfriendlyName | ForEach-Object {[System.Text.Encoding]::ASCII.GetString($_.UserFriendlyName)}")[0]

                sheet['B14'].value = str(data['display_name'].replace('\x00', ''))#.encode("ascii",errors="ignore")
            except:
                pass

            try:
                data['display_diagonal'] = shellCommand('Get-WmiObject -computername ' + data['pc_name'] + ' -Namespace root\wmi -Class WmiMonitorBasicDisplayParams | select @{N="Size"; E={[System.Math]::Round(([System.Math]::Sqrt([System.Math]::Pow($_.MaxHorizontalImageSize, 2) + [System.Math]::Pow($_.MaxVerticalImageSize, 2))/2.54),2)} }')[3].replace('\r', '') + '\'\''

                sheet['D17'].value = data['display_diagonal']
            except:
                pass

            try:
                stdout = shellCommand('Get-WmiObject -Computername ' + data['pc_name'] + ' -Class MSFT_PhysicalDisk -Namespace root\Microsoft\Windows\Storage| select MediaType, size')
                data['diskdrive_hdd'] = '-'
                data['diskdrive_ssd'] = '-'
                data['diskdrive_size'] = 0

                for i in stdout[3: len(stdout) - 3]:
                    tmp = sParse(i.replace('\r', '')).split(' ')
                    if (tmp[1] == 'HDD') or (tmp[1] == '3'):
                        data['diskdrive_hdd'] = '+'
                        data['diskdrive_size'] += int(tmp[2])
                    if (tmp[1] == 'SSD') or (tmp[1] == '4'):
                        data['diskdrive_ssd'] = '+'
                        data['diskdrive_size'] += int(tmp[2])

                data['diskdrive_size'] = str(round(data['diskdrive_size'] / 1024 / 1024 / 1024, 2)) + ' Gb'

                sheet['I11'].value = data['diskdrive_size']
                sheet['E11'].value = data['diskdrive_hdd']
                sheet['G11'].value = data['diskdrive_ssd']

            except:
                stdout, stderr, rc = c.run_executable("cmd.exe",arguments="/c wmic logicaldisk get drivetype, size")
                data['diskdrive_size'] = 0

                for i in cmdParse(stdout)[1: len(cmdParse(stdout))]:
                    tmp = sParse(i.replace('\r', '')).split(' ')
                    if tmp[0] == '3':
                        data['diskdrive_size'] += int(tmp[1])

                data['diskdrive_size'] = str(round(data['diskdrive_size'] / 1024 / 1024 / 1024, 2)) + ' Gb'

                sheet['I11'].value = data['diskdrive_size']


            print('\n*Пользователь ' + currentUser + ' - успешно')

        except:
            print('\n*Пользователь ' + currentUser + ' - неудачно')
            failedUsers.append(currentUser.split(' ')[0] + ', ' + data['pc_name'] + ', ' + data['position'])

        finally:
            c.disconnect()
            wb.save(upravlenie + '/' + currentUser + '.xlsx')
            print('#' * 62)
            output_rep(data)
            print('#' * 62)

    data.clear()
    userList.clear()

    f = open('failed.txt', 'a')

    if len(failedUsers) > 0:
        print('Неудачно: ')
        for i in failedUsers:
            f.write(i + '\n')
            print(i)
    else:
        print('Созданы таблицы всех указанных пользователей')

    f.write('\n\n')
    f.close()
    failedUsers.clear()
